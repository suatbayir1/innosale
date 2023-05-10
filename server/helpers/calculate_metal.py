from openpyxl import load_workbook
import os

def get_all_values(filename_):
    wb = load_workbook('/root/innosale/server/static/ParcaListesi.xlsx')
    sheet_obj = wb.active

    filename_ = filename_.split(".")
    filename_ = filename_[0]
    for i in range(2,895):

        cell_obj = sheet_obj.cell(row = i, column = 1)
        teklifno = str(cell_obj.value)
        m = teklifno.split("-")
        temp = m[0]
        t = "_"
        temp += t
        temp += m[1]

        n = "_Teklifid_"
        temp += n
        
        count = 0

        cell_obj2 = sheet_obj.cell(row = i, column = 3)
        teklifid = str(cell_obj2.value)
        temp += teklifid
        file_name = temp.split(".")
        filename = file_name[0]
        
        if filename != filename_:
            continue

        print(filename)
        
        #sacKalınlık
        cell_obj3 = sheet_obj.cell(row= i , column = 4)
        sac_z = float(cell_obj3.value)


        #NetX
        cell_obj4 = sheet_obj.cell(row=i, column=6)
        sac_x = float(cell_obj4.value)

        #NetY
        cell_obj7 = sheet_obj.cell(row=i, column=7)
        sac_y = float(cell_obj7.value)
        
        #KonturBoyu
        cell_obj8 = sheet_obj.cell(row=i, column=8)
        kontur_boyu = float(cell_obj8.value)
        
        
        #Acınım_yuzeyAlanı
        cell_obj9 = sheet_obj.cell(row=i, column=9)
        yuzey_alani = float(cell_obj9.value)
        
        
        
        #SacTsMax
        cell_obj10 = sheet_obj.cell(row=i, column=10)
        sacTsMax = float(cell_obj10.value)
        
        
        #SacUzama
        cell_obj11 = sheet_obj.cell(row=i, column=11)
        sacUzama = float(cell_obj11.value)
        
        
        
        dic_parameters = {
        "sac": sac_z,
        "acinim_x": sac_x,
        "acinim_y": sac_y,
        "kontur": kontur_boyu,
        "alan": yuzey_alani,
        "SacTsMax": sacTsMax,
        "SacUzama": sacUzama
        
        }
        print(dic_parameters)

        return dic_parameters
        


def get_values(filename_):
    wb = load_workbook('/root/innosale/server/static/ParcaListesi.xlsx')
    sheet_obj = wb.active

    filename_ = filename_.split(".")
    filename_ = filename_[0]
    
    for i in range(2,895):

        cell_obj = sheet_obj.cell(row = i, column = 1)
        teklifno = str(cell_obj.value)
        m = teklifno.split("-")
        temp = m[0]
        t = "_"
        temp += t
        temp += m[1]

        n = "_Teklifid_"
        temp += n
        
        count = 0

        cell_obj2 = sheet_obj.cell(row = i, column = 3)
        teklifid = str(cell_obj2.value)
        temp += teklifid
        file_name = temp.split(".")
        filename = file_name[0]
        
        if filename != filename_:
            continue

        print(filename)
        
        #sacKalınlık
        cell_obj3 = sheet_obj.cell(row= i , column = 4)
        sac_z = float(cell_obj3.value)


        #NetX
        cell_obj4 = sheet_obj.cell(row=i, column=6)
        sac_x = float(cell_obj4.value)

        #NetY
        cell_obj5 = sheet_obj.cell(row=i, column=7)
        sac_y = float(cell_obj5.value)
        
        #KonturBoyu
        cell_obj5 = sheet_obj.cell(row=i, column=8)
        kontur_boyu = float(cell_obj5.value)
        

        
        #Acınım_yuzeyAlanı
        cell_obj5 = sheet_obj.cell(row=i, column=9)
        yuzey_alani = float(cell_obj5.value)
        
        
        
        dic_parameters = {
        "sac": sac_z,
        "acinim_x": sac_x,
        "acinim_y": sac_y,
        "kontur": kontur_boyu,
        "alan": yuzey_alani
        }
        print(dic_parameters)
        
        return dic_parameters
        
        
def f(sackalinlik,netx,nety,konturboyu,yuzeyalani):
    wb = load_workbook('/root/innosale/server/static/ParcaListesi.xlsx')
    
    sheet_obj = wb.active
    lis = []
    x = 0
        
    def fn2():
        file_list2 = os.listdir("/root/innosale/server/static/convertedfiles2")
        return file_list2

    filey = []
    file2 = fn2()
    for i in file2:
        y = i.split(".")
        filey.append(y[0])

    
    for i in range(2,895):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        teklifno = str(cell_obj.value)
        m = teklifno.split("-")
        temp=m[0]
        t = "_"
        temp += t
        temp += m[1]

        n = "_Teklifid_"
        temp += n
        
        count = 0

        cell_obj2 = sheet_obj.cell(row = i, column = 3)
        teklifid = str(cell_obj2.value)
        temp += teklifid
        file_name = temp.split(".")
        filename = file_name[0]

        if filename not in filey:
            continue

        #sacKalınlık
        cell_obj3 = sheet_obj.cell(row= i , column = 4)
        sac_z = float(cell_obj3.value)

        if sac_z < sackalinlik[1] and sac_z > sackalinlik[0]:
            count = count + 1

        #NetX
        cell_obj4 = sheet_obj.cell(row=i, column=6)
        sac_x = float(cell_obj4.value)

        if sac_x < netx[1] and sac_x > netx[0]:
            count = count + 1        

        #NetY
        cell_obj5 = sheet_obj.cell(row=i, column=7)
        sac_y = float(cell_obj5.value)

        if sac_y < nety[1] and sac_y > nety[0]:
            count = count + 1
        
        #KonturBoyu
        cell_obj5 = sheet_obj.cell(row=i, column=8)
        kontur_boyu = float(cell_obj5.value)
        
        if kontur_boyu < konturboyu[1] and kontur_boyu > konturboyu[0]:
            count = count + 1
        
        #Acınım_yuzeyAlanı
        cell_obj5 = sheet_obj.cell(row=i, column=9)
        yuzey_alani = float(cell_obj5.value)
        
        if yuzey_alani < yuzeyalani[1] and yuzey_alani > yuzeyalani[0]:
            count = count + 1
        

        if count == 5 :
            lis.append(filename)
                        
   
    #print("----------------------------------------------------------------------------")
    #print("----------------------------------------------------------------------------")
    #print("----------------------------------------------------------------------------")
    print(lis)
    return lis








